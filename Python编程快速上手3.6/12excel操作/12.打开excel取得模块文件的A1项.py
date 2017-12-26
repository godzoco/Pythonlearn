#先安装openyxl 模块 pip install openyxl
import openpyxl
import os
w = openpyxl.load_workbook('example.xlsx')
s = w.get_sheet_by_name('Sheet1')

print(s['A1'].value)

print(s['B3'].value)
print(s.cell(row=1 ,column=2).value)#row是行
for i in range(1,7):
    print(i,s.cell(row=i,column=2).value)
print('8个弄完7个')
for i in range(1,8):
    print(i,s.cell(row=i,column=2).value)

#print(s.get_highest_row())
#print(s.get_highest_column())
